from __future__ import annotations
from pathlib import Path
from typing import Any
from openpyxl import load_workbook
from ..database.repositories import Repository
from ..database.schema import connect


def norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def keyify(value: Any) -> str:
    return norm(value).lower().replace("/", " ").replace("-", " ").replace("_", " ").strip()


def as_int(value: Any, default: int = 0) -> int:
    try:
        if value in (None, ""):
            return default
        text = str(value).replace(",", "").strip()
        if text in ("—", "-"):
            return default
        return int(float(text))
    except Exception:
        return default


def find_header(ws, expected_terms: set[str], max_scan_rows: int = 8) -> tuple[int, dict[str, int]]:
    """Find a likely header row in the first few rows.

    The source workbook often has a title row and/or description row before the real
    column names. This function scans for a row containing expected column labels.
    """
    best_row = 1
    best_score = -1
    best_headers: dict[str, int] = {}
    for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        headers: dict[str, int] = {}
        score = 0
        for idx, value in enumerate(row, start=0):
            k = keyify(value)
            if not k:
                continue
            headers[k] = idx
            if k in expected_terms:
                score += 10
            elif len(k) <= 60 and any(term in k for term in expected_terms):
                score += 1
        if score > best_score:
            best_row = row_index
            best_score = score
            best_headers = headers
    return best_row, best_headers


def get(row, headers, *names, default=""):
    for name in names:
        target = keyify(name)
        # exact
        if target in headers and headers[target] < len(row):
            v = row[headers[target]]
            if v is not None:
                return v
        # contains fallback, e.g. armor class text -> armor class
        for key, idx in headers.items():
            if (target and (target == key or target in key or key in target)) and idx < len(row):
                v = row[idx]
                if v is not None:
                    return v
    return default


class SpreadsheetImporter:
    def __init__(self, db_path: Path):
        self.repo = Repository(db_path)

    def import_file(self, path: Path) -> int:
        wb = load_workbook(path, data_only=True, read_only=True)
        total = 0
        try:
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lower = sheet.lower()
                if "monster database" in lower or lower == "monsters" or "monster" in lower:
                    total += self._import_monsters(ws)
                elif lower in {"players", "player build helper"} or "player character" in lower:
                    total += self._import_players(ws)
                elif "weapons database" in lower or "weapon" in lower:
                    total += self._import_weapons(ws)
                elif "armor database" in lower or "armour" in lower or lower == "armor":
                    total += self._import_armor(ws)
                elif "equipment database" in lower or lower == "equipment" or "gear" in lower:
                    total += self._import_equipment(ws)
                elif "magic item" in lower:
                    total += self._import_magic_items(ws)
                elif "spell lists" in lower:
                    total += self._import_rules(ws, sheet)
                elif "spell" in lower and "tracker" not in lower:
                    total += self._import_spells(ws)
                elif any(x in lower for x in ["condition", "feat", "class", "race", "species", "background"]):
                    total += self._import_rules(ws, sheet)
        finally:
            wb.close()
        self.repo.add_import_history(str(path), total, "spreadsheet import")
        return total

    def import_rules_only(self, path: Path) -> int:
        """Refresh SRD character-building references without touching player data."""
        wb = load_workbook(path, data_only=True, read_only=True)
        total = 0
        try:
            for sheet in wb.sheetnames:
                lower = sheet.lower()
                if any(x in lower for x in ["condition", "feat", "class", "race", "species", "background"]):
                    total += self._import_rules(wb[sheet], sheet)
        finally:
            wb.close()
        self.repo.add_import_history(str(path), total, "SRD rules refresh")
        return total

    def _iter_data(self, ws, expected: set[str]):
        header_row, headers = find_header(ws, expected)
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or all(v is None or norm(v) == "" for v in row):
                continue
            yield row, headers

    def _import_players(self, ws) -> int:
        count = 0
        expected = {"name", "player name", "character name", "species", "class", "level"}
        for row, headers in self._iter_data(ws, expected):
            name = norm(get(row, headers, "character name", "name", "player name", "player"))
            if not name or name.lower().startswith("player character database"):
                continue
            max_hp = as_int(get(row, headers, "max hp", "hp", "hit points"), 1)
            self.repo.upsert_player({
                "name": name,
                "species": norm(get(row, headers, "species", "race")),
                "class_name": norm(get(row, headers, "class", "class name")),
                "subclass": norm(get(row, headers, "subclass")),
                "background": norm(get(row, headers, "background")),
                "level": as_int(get(row, headers, "level"), 1),
                "armor_class": as_int(get(row, headers, "armor class", "ac"), 10),
                "max_hp": max_hp,
                "current_hp": as_int(get(row, headers, "current hp", "hp"), max_hp),
                "initiative_mod": as_int(get(row, headers, "initiative mod", "initiative", "dex mod"), 0),
                "notes": norm(get(row, headers, "notes")),
            })
            count += 1
        return count

    def _import_monsters(self, ws) -> int:
        count = 0
        expected = {"monster name", "name", "cr", "xp", "ac", "max hp", "type class"}
        for row, headers in self._iter_data(ws, expected):
            name = norm(get(row, headers, "monster name", "name", "monster"))
            if not name or name.lower().startswith("monster database"):
                continue
            self.repo.upsert_monster({
                "name": name,
                "size": norm(get(row, headers, "size")),
                "type": norm(get(row, headers, "type class", "type", "class")),
                "alignment": norm(get(row, headers, "alignment")),
                "armor_class": as_int(get(row, headers, "ac", "armor class"), 10),
                "hit_points": as_int(get(row, headers, "max hp", "hit points", "hp"), 1),
                "speed": norm(get(row, headers, "speed")),
                "challenge_rating": norm(get(row, headers, "cr", "challenge rating")),
                "xp": as_int(get(row, headers, "xp"), 0),
                "str_score": as_int(get(row, headers, "str", "strength"), 0),
                "dex_score": as_int(get(row, headers, "dex", "dexterity"), 0),
                "con_score": as_int(get(row, headers, "con", "constitution"), 0),
                "int_score": as_int(get(row, headers, "int", "intelligence"), 0),
                "wis_score": as_int(get(row, headers, "wis", "wisdom"), 0),
                "cha_score": as_int(get(row, headers, "cha", "charisma"), 0),
                "source": norm(get(row, headers, "source")) or "spreadsheet",
                "notes": norm(get(row, headers, "notes", "description")),
            })
            count += 1
        return count

    def _import_weapons(self, ws) -> int:
        count = 0
        expected = {"weapon name", "category", "damage dice", "damage type", "properties"}
        for row, headers in self._iter_data(ws, expected):
            name = norm(get(row, headers, "weapon name", "name"))
            if not name or name.lower().startswith("weapons database"):
                continue
            damage = norm(get(row, headers, "damage dice", "damage"))
            dtype = norm(get(row, headers, "damage type"))
            if dtype:
                damage = f"{damage} {dtype}".strip()
            self.repo.upsert_reference("weapons", {
                "name": name,
                "category": norm(get(row, headers, "category")),
                "damage": damage,
                "properties": norm(get(row, headers, "properties")),
                "weight": norm(get(row, headers, "weight")),
                "cost": norm(get(row, headers, "cost")),
                "notes": norm(get(row, headers, "source")),
            })
            count += 1
        return count

    def _import_armor(self, ws) -> int:
        count = 0
        expected = {"armor name", "category", "armor class text", "base ac"}
        for row, headers in self._iter_data(ws, expected):
            name = norm(get(row, headers, "armor name", "name"))
            if not name or name.lower().startswith("armor database"):
                continue
            self.repo.upsert_reference("armor", {
                "name": name,
                "category": norm(get(row, headers, "category")),
                "armor_class": norm(get(row, headers, "armor class text", "armor class", "base ac", "ac")),
                "strength": norm(get(row, headers, "strength required", "strength")),
                "stealth": norm(get(row, headers, "stealth")),
                "weight": norm(get(row, headers, "weight")),
                "cost": norm(get(row, headers, "cost")),
                "notes": norm(get(row, headers, "source")),
            })
            count += 1
        return count

    def _import_equipment(self, ws) -> int:
        count = 0
        expected = {"item name", "category", "weight", "cost"}
        for row, headers in self._iter_data(ws, expected):
            name = norm(get(row, headers, "item name", "equipment name", "name"))
            if not name or name.lower().startswith("equipment database"):
                continue
            self.repo.upsert_reference("equipment", {
                "name": name,
                "category": norm(get(row, headers, "category")),
                "cost": norm(get(row, headers, "cost")),
                "weight": norm(get(row, headers, "weight")),
                "notes": norm(get(row, headers, "combat use", "source", "notes")),
            })
            count += 1
        return count

    def _import_magic_items(self, ws) -> int:
        count = 0
        expected = {"magic item name", "item type", "rarity", "requires attunement"}
        for row, headers in self._iter_data(ws, expected):
            name = norm(get(row, headers, "magic item name", "item name", "name"))
            if not name or name.lower().startswith("magic item database"):
                continue
            self.repo.upsert_reference("magic_items", {
                "name": name,
                "rarity": norm(get(row, headers, "rarity")),
                "item_type": norm(get(row, headers, "item type", "type")),
                "attunement": norm(get(row, headers, "requires attunement", "attunement")),
                "notes": norm(get(row, headers, "srd type line", "passive bonus rule", "source", "notes")),
            })
            count += 1
        return count

    def _import_spells(self, ws) -> int:
        count = 0
        expected = {"spell", "level", "school", "casting time", "range"}
        for row, headers in self._iter_data(ws, expected):
            name = norm(get(row, headers, "spell", "spell name", "name"))
            if not name or name.lower().startswith("player db"):
                continue
            self.repo.upsert_reference("spells", {
                "name": name,
                "level": norm(get(row, headers, "level")),
                "school": norm(get(row, headers, "school")),
                "casting_time": norm(get(row, headers, "casting time")),
                "range_text": norm(get(row, headers, "range")),
                "components": norm(get(row, headers, "components")),
                "duration": norm(get(row, headers, "duration")),
                "description": norm(get(row, headers, "description")),
            })
            count += 1
        return count

    def _import_rules(self, ws, category: str) -> int:
        count = 0
        expected = {"name", "species", "background", "class", "subclass", "feat", "ability", "condition"}
        for row, headers in self._iter_data(ws, expected):
            name = (
                norm(get(row, headers, "name")) or norm(get(row, headers, "species")) or
                norm(get(row, headers, "background")) or norm(get(row, headers, "class")) or
                norm(get(row, headers, "subclass")) or norm(get(row, headers, "feat")) or
                norm(get(row, headers, "ability")) or norm(get(row, headers, "condition"))
            )
            if not name or name.lower().startswith("srd"):
                continue
            desc = norm(get(row, headers, "description", "summary", "trait summary", "feature summary", "notes", default=""))
            with connect(self.repo.db_path) as conn:
                conn.execute(
                    "INSERT OR IGNORE INTO rules_reference(category,name,description) VALUES(?,?,?)",
                    (category, name, desc),
                )
            count += 1
        return count
